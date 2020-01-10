'''
Created on 25 Jul 2019

@author: daim
'''


from etilog.models import (Country, SustainabilityDomain, SustainabilityTag, 
                           SustainabilityTendency, ImpactEvent,
                           Company, SubsidiaryOwner, SupplierRecipient
                           )
from openpyxl import load_workbook
import warnings
import os


def parse_xcl(): #, eventlist, eventtype_dict):   
    base_dir = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
    spath = os.path.join(base_dir, 'DB_ImportFields.xlsx')
    print (spath)
   
    warnings.simplefilter("ignore")
    
    wb = load_workbook(spath )#, read_only = True)#read_only = True) 

    warnings.simplefilter("default")
    
    parse_companies_relations()
    
    #imp_originalmodel(wb)
    
    
    #not needed anymore as already parsed
    #parse_tags_domain_tendency() #after parse tags
    #parse_ies_domain_tendency()

def imp_originalmodel(wb):
    parse_country(wb)
    
    parse_domain(wb)
    
    parse_tags(wb) #after sustcateg and domains
    

def parse_country(wb):
    
    sheet = wb['Country']
    rowcount = sheet.max_row 


            
    for cell in sheet [1]:
        col_strs = ['Name', '2alpha','3alpha', 'Nr']
        col_dict = {}
        for col_str in col_strs:   
            for cell in sheet [1]:               
                if cell.value ==  col_str:
                    col_dict[col_str] = cell.column

    i = 3 #without header
   
    while i < rowcount + 1: #sheet_iter = sheet.iter_rows
        
        name_str = sheet.cell(row = i, column = col_dict['Name']).value
        num_int = int(sheet.cell(row = i, column = col_dict['Nr']).value)
        a2_str = sheet.cell(row = i, column = col_dict['2alpha']).value
        a3_str = sheet.cell(row = i, column = col_dict['3alpha']).value
        
        default_data={'name': name_str, 
                      'alpha2code' : a2_str,
                      'alpha3code' : a3_str
                      }
        
        country , created = Country.objects.update_or_create(numeric = num_int,    
                                                             defaults = default_data)
        country.save()
        i += 1
        
def parse_domain(wb):
    
    sheet = wb['SustDomain']
    rowcount = sheet.max_row 


            
    for cell in sheet [1]:
        col_strs = ['Nr', 'Name']
        col_dict = {}
        for col_str in col_strs:   
            for cell in sheet [1]:               
                if cell.value ==  col_str:
                    col_dict[col_str] = cell.column

    i = 3 #without header
   
    while i < rowcount + 1: #sheet_iter = sheet.iter_rows
        
        name_str = sheet.cell(row = i, column = col_dict['Name']).value
        num_int = int(sheet.cell(row = i, column = col_dict['Nr']).value)
        
        

        default_data={'name': name_str, 
                      }
        
        domain_db , created = SustainabilityDomain.objects.update_or_create(impnr = num_int,
                                                                            defaults = default_data)
        domain_db.save()         
        i += 1 



def parse_tags(wb):
    
    sheet = wb['Tags']
    rowcount = sheet.max_row 


            
    for cell in sheet [1]:
        col_strs = ['Nr', 'Name','CatID', 'description']
        col_dict = {}
        for col_str in col_strs:   
            for cell in sheet [1]:               
                if cell.value ==  col_str:
                    col_dict[col_str] = cell.column

    i = 3 #without header
   
    while i < rowcount + 1: #sheet_iter = sheet.iter_rows
        
        name_str = sheet.cell(row = i, column = col_dict['Name']).value
        print (sheet.cell(row = i, column = col_dict['Nr']).value)
        num_int = int(sheet.cell(row = i, column = col_dict['Nr']).value)
        categ_str = sheet.cell(row = i, column = col_dict['CatID']).value
        descript_str = sheet.cell(row = i, column = col_dict['description']).value

        default_data={'name': name_str, 
                      'description' : descript_str  
                      }
        
        tag_db , created = SustainabilityTag.objects.update_or_create(impnr = num_int,
                                                                      defaults = default_data)
        tag_db.save() #needed for manytomany
        
        if categ_str:
            cat_list = categ_str.split(';')
            catdb_list = []
            for categ_int in cat_list:
                catdb = SustainabilityCategory.objects.get(impnr = int(categ_int) )
                catdb_list.append(catdb)
                
            tag_db.sust_categories.set(catdb_list)
            tag_db.save() 
            
        i += 1 
        
def parse_tags_domain_tendency():
    q = SustainabilityTag.objects.all()
    for el_db in q:
        q_sust = el_db.sust_categories.all()
        ten_list = []
        dom_list = []
        for st_db in q_sust:
            dom_db, ten_db = get_domain_tendency(st_db)
            dom_list.append(dom_db)
        el_db.sust_domains.set(dom_list)
        el_db.sust_tendency = ten_db
        el_db.save()

def parse_ies_domain_tendency():
    q = ImpactEvent.objects.all()
    for el_db in q:
        st_db = el_db.sust_category
        dom_db, ten_db = get_domain_tendency(st_db)
        el_db.sust_domain = dom_db
        el_db.sust_tendency = ten_db
        el_db.save()
            
        
def get_domain_tendency(st_db):
    dom_db =  st_db.sust_domain
    sname = st_db.name
    pos = 'positive'
    neg = 'negative' 
    contr = 'controversial' 
    if pos[:4] in sname:
        name_str = pos
        def_data = {'name': name_str }
        ten_db, created = SustainabilityTendency.objects.get_or_create(name__icontains = name_str,
                                                defaults = def_data)
        
        
     
      
    elif neg[:4] in sname:
        name_str = neg
        def_data = {'name': name_str }
        ten_db, created = SustainabilityTendency.objects.get_or_create(name__icontains = name_str,
                                               defaults = def_data)
 
    elif contr[:4] in sname:
        name_str = contr
        def_data = {'name': name_str }
        ten_db, created = SustainabilityTendency.objects.get_or_create(name__icontains = name_str,
                                              defaults = def_data)
    
    ten_db.save()
    
    return dom_db, ten_db
    
def parse_companies_relations():
    def create_relobj(q, cls):
        wrong_relations = []
        for obj in q.order_by('pk'): #all relations are doubled, first is correct one
            objid = obj.pk
            if objid in wrong_relations:
                continue
            
            fcomp= obj.from_company
            tcomp = obj.to_company
            #get wrong ones where its the opposite way with higher pk
            wrong_obj = q.get(from_company = tcomp, to_company = fcomp)
            wrong_relations.append(wrong_obj.pk)
            
            #create if not exist
            if cls == 'sub':              
                #target is a subsidiary
                SubsidiaryOwner.objects.update_or_create(owner_company = fcomp, 
                                                         subsidiary_company = tcomp)
            
            elif cls == 'owner':
                #target is a owner
                SubsidiaryOwner.objects.update_or_create(owner_company = tcomp, 
                                                         subsidiary_company = fcomp)
            
            elif cls == 'suppliers':              
                #target is a supplier
                SupplierRecipient.objects.update_or_create(recipient_company = fcomp, 
                                                           supplier_company = tcomp)
            
            elif cls == 'recipients':              
                #target is a recipient
                SupplierRecipient.objects.update_or_create(recipient_company = tcomp, 
                                                           supplier_company = fcomp)
            
            
    
    co1 = Company.objects.first()
    
    #get all subsidiary of through model
    subs = co1.subsidiary_old.through.objects.all()
    create_relobj(subs, 'sub')
    
    owners = co1.owner_old.through.objects.all()
    create_relobj(owners, 'owner')
    
    owners = co1.supplier_old.through.objects.all()
    create_relobj(owners, 'suppliers')
    
    owners = co1.recipient_old.through.objects.all()
    create_relobj(owners, 'recipients')